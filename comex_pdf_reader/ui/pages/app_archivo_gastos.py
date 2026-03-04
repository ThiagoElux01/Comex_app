# -*- coding: utf-8 -*-
import re
import numpy as np
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from pandas.api.types import is_numeric_dtype

# -----------------------------------------------------------------------------
# Estado e helpers
# -----------------------------------------------------------------------------

def _ensure_state():
    """
    Garante que todas as chaves necessárias existam em st.session_state,
    mesmo se já houver um dicionário 'aag_state' antigo/incompleto na sessão.
    """
    # Cria o dicionário de estado principal se não existir
    if "aag_state" not in st.session_state or not isinstance(st.session_state["aag_state"], dict):
        st.session_state["aag_state"] = {}
    aag = st.session_state["aag_state"]

    # Keys dos uploaders separadas por modo (evita conflito de cache do Streamlit)
    aag.setdefault("uploader_key_estado", "aag_estado_upl_1")
    aag.setdefault("uploader_key_pg", "aag_pg_upl_1")
    aag.setdefault("uploader_key_cuenta", "aag_cuenta_upl_1")

    # Última ação (reserva)
    aag.setdefault("last_action", None)

    # Modo atual da página
    if "aag_mode" not in st.session_state:
        st.session_state["aag_mode"] = "estado"  # default


def _set_mode(mode: str):
    st.session_state["aag_mode"] = mode


# ==== Helpers de formatação para 'Chave' ====
def _remove_newlines(s: pd.Series) -> pd.Series:
    """Remove \r e \n de uma Series de strings."""
    return (
        s.astype(str)
         .str.replace(r"[\r\n]+", " ", regex=True)  # troca por espaço (ou use "" se quiser colar)
         .str.strip()
    )
def _fmt_date_ddmmyyyy(value) -> str:
    """Converte vários tipos de data para 'dd/mm/aaaa' como string."""
    if pd.isna(value):
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce")  # formato ISO já é entendido
        if pd.isna(dt):
            return ""
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return ""


def _fmt_num_2dec_point(value) -> str:
    """Formata número com 2 casas, ponto como decimal, sem milhares (ex.: 1234.50)."""
    try:
        f = float(value)
        return f"{f:.2f}"
    except Exception:
        return ""


def _str_or_empty(x) -> str:
    return "" if x is None or (isinstance(x, float) and np.isnan(x)) else str(x).strip()


def _fmt_transno_keep_zeros(x, width: int = 9) -> str:
    """
    Normaliza 'TransactionNo'/'Transacción' para apenas dígitos e zera à esquerda
    até 'width' casas. Ex.: 18528 -> 000018528 (width=9).
    Trata corretamente leituras como 18528.0 (float) e strings com texto.
    """
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    if isinstance(x, (int, np.integer)):
        s = str(int(x))
    elif isinstance(x, (float, np.floating)):
        s = str(int(round(float(x))))
    else:
        s = str(x).strip()
        s_norm = s.replace(",", ".")
        if re.fullmatch(r"\d+\.\d+", s_norm):
            try:
                s = str(int(float(s_norm)))
            except Exception:
                s = re.sub(r"\D", "", s)
        else:
            s = re.sub(r"\D", "", s)
    if not s:
        return ""
    return s.zfill(width)


# ==== Helpers robustos para datas na PLANTILLA ====
_EXCEL_ORIGIN = "1899-12-30"  # origem do Excel (Windows)


def _to_datetime_from_mixed_excel_and_strings(s: pd.Series) -> pd.Series:
    """
    Converte coluna de datas da Plantilla de Gastos para dtype datetime64[ns],
    tratando 3 casos:
      1) seriais Excel (número puro ou string numérica) -> origin 1899-12-30
      2) strings ISO (ex.: '2026-01-02 00:00:00') -> pd.to_datetime padrão
      3) outras strings -> tentativa genérica de parse
    """
    # 1) Tenta converter seriais (numérico)
    s_str = s.astype(str).str.strip()
    is_num_like = s_str.str.match(r"^\d+(\.\d+)?$").fillna(False)
    dt = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")

    if is_num_like.any():
        s_num = pd.to_numeric(s_str[is_num_like], errors="coerce")
        dt.loc[is_num_like] = pd.to_datetime(s_num, unit="d", origin=_EXCEL_ORIGIN, errors="coerce")

    # 2) Para o restante, remove o sufixo ' HH:MM:SS' se existir e faz parse padrão
    rest = ~is_num_like
    if rest.any():
        s_rem = s_str[rest].str.replace(r"\s+\d{2}:\d{2}:\d{2}$", "", regex=True)
        # Primeiro passe: pd.to_datetime padrão (ISO funciona direto)
        dt1 = pd.to_datetime(s_rem, errors="coerce")
        dt.loc[rest] = dt1

        # Fallback (pouco provável aqui): tenta mais uma vez
        still = dt1.isna()
        if still.any():
            s2 = s_rem[still]
            dt2 = pd.to_datetime(s2, errors="coerce")
            dt.loc[s2.index] = dt2

    return dt


def _fmt_date_series_ddmmyyyy(s: pd.Series) -> pd.Series:
    """Formata uma Series de datas em 'dd/mm/yyyy' como texto para compor a Chave."""
    s_dt = pd.to_datetime(s, errors="coerce")
    return s_dt.dt.strftime("%d/%m/%Y").fillna("")


# -----------------------------------------------------------------------------
# Limpieza Plantilla Gastos — helper robusto
# -----------------------------------------------------------------------------

def limpiar_plantilla_contra_cuenta(
    df_pg: pd.DataFrame,
    df_cuenta: pd.DataFrame,
    chave_col: str = "Chave",
    tol_soma: float = 0.005
) -> tuple[pd.DataFrame, dict]:
    """
    Remove excedentes na Plantilla por 'Chave' para igualar a contagem às ocorrências no GL0061.
    - Só afeta chaves presentes na Cuenta.
    - Mantém a ordem original.
    - Guard: se contagem e soma por chave batem (dentro da tolerância), não remove nada.
    """
    if chave_col not in df_pg.columns:
        raise ValueError("Plantilla de Gastos não contém a coluna 'Chave'. Rode a etapa da Plantilla antes.")
    if chave_col not in df_cuenta.columns:
        raise ValueError("Cuenta (GL0061) não contém a coluna 'Chave'. Rode a etapa de Cuenta antes.")

    def _norm_key_series(s: pd.Series) -> pd.Series:
        return (
            s.astype(str)
             .str.replace("\u2212", "-", regex=False)
             .str.replace("\xa0", " ", regex=False)
             .str.replace(r"\s+", " ", regex=True)
             .str.strip()
        )

    df_pg = df_pg.copy()
    df_cuenta = df_cuenta.copy()
    df_pg["_key_norm"] = _norm_key_series(df_pg[chave_col])
    df_cuenta["_key_norm"] = _norm_key_series(df_cuenta[chave_col])

    first_original_key = df_pg.groupby("_key_norm")[chave_col].first()

    cnt_pg = df_pg["_key_norm"].value_counts()
    cnt_cuenta = df_cuenta["_key_norm"].value_counts()

    amount_col = None
    for c in df_pg.columns:
        cl = str(c).strip().lower()
        if cl == "amount" or "amount" in cl:
            amount_col = c
            break
    cuenta_amt_col = "Saldo Real" if "Saldo Real" in df_cuenta.columns else None

    sum_pg = pd.Series(dtype=float)
    sum_cuenta = pd.Series(dtype=float)
    if amount_col is not None:
        sum_pg = pd.to_numeric(df_pg[amount_col], errors="coerce").groupby(df_pg["_key_norm"]).sum(min_count=1)
    if cuenta_amt_col is not None:
        sum_cuenta = pd.to_numeric(df_cuenta[cuenta_amt_col], errors="coerce").groupby(df_cuenta["_key_norm"]).sum(min_count=1)

    keep_limit_map = cnt_cuenta.to_dict()

    rank = df_pg.groupby("_key_norm").cumcount() + 1
    keep_limit = df_pg["_key_norm"].map(keep_limit_map)

    mask_keep = keep_limit.isna() | (rank <= keep_limit.fillna(np.inf))

    if not sum_pg.empty and not sum_cuenta.empty:
        keys_ok = []
        for k, cpg in cnt_pg.items():
            cct = int(cnt_cuenta.get(k, 0))
            if cpg == cct and cct > 0:
                spg = float(sum_pg.get(k, 0.0))
                scu = float(sum_cuenta.get(k, 0.0))
                if abs(spg - scu) <= tol_soma:
                    keys_ok.append(k)
        if keys_ok:
            mask_keep = mask_keep | df_pg["_key_norm"].isin(keys_ok)

    df_clean = df_pg[mask_keep].copy().reset_index(drop=True)
    df_clean.drop(columns=["_key_norm"], inplace=True, errors="ignore")

    removed_total = int((~mask_keep).sum())
    removed_by_key = {}
    keys_with_drop = []
    for k, cpg in cnt_pg.items():
        cct = int(cnt_cuenta.get(k, 0))
        if cpg > cct and cct > 0:
            diff = cpg - cct
            k_disp = str(first_original_key.get(k, k))
            removed_by_key[k_disp] = diff
            keys_with_drop.append(k)

    stats = {
        "rows_original": int(len(df_pg)),
        "rows_clean": int(len(df_clean)),
        "rows_removed": removed_total,
        "keys_with_removal": len(keys_with_drop),
        "removed_by_key": dict(sorted(removed_by_key.items(), key=lambda kv: kv[1], reverse=True)),
    }
    return df_clean, stats


# -----------------------------------------------------------------------------
# Parsers - ESTADO DE CUENTA (.txt)
# -----------------------------------------------------------------------------
_NUM = r"(\-?\d[\d,]*\.\d{2}\-?)"  # número com milhares e 2 decimais; pode terminar com '-' (negativo)


def _clean_num(s: str) -> float | None:
    """Converte strings como '12,345.67-' em float (negativo)."""
    if s is None:
        return None
    s = str(s).strip()
    if s == "":
        return None
    neg = s.endswith("-")
    s = s[:-1] if neg else s
    s = s.replace(",", "")
    try:
        v = float(s)
        return -v if neg else v
    except Exception:
        return None


def parse_estado_cuenta_txt(texto: str) -> pd.DataFrame:
    """
    Lê um relatório 'Listado de Saldos' em texto e retorna um DataFrame com:
    ['CTA','Descripción','Sal OB','Saldo OB','Período','Saldo CB']
    """
    linhas = texto.splitlines()

    start_idx = 0
    for i, ln in enumerate(linhas):
        if "CTA" in ln and "Descripci" in ln:
            start_idx = i + 1
            break

    dados = []
    tail_re = re.compile(rf"\s*{_NUM}\s+{_NUM}\s+{_NUM}\s+{_NUM}\s*$")

    for ln in linhas[start_idx:]:
        raw = ln.rstrip()
        if not raw:
            continue
        if set(raw.strip()) in [{"="}, {"-"}] or "Scala" in raw or "Electrolux" in raw:
            continue

        m = tail_re.search(raw)
        if not m:
            continue

        left = raw[: m.start()].rstrip()
        if not left:
            continue

        parts = left.split()
        cta = parts[0] if parts else ""
        descr = left[len(cta):].strip() if parts else left.strip()

        sal_ob, saldo_ob, periodo, saldo_cb = (_clean_num(x) for x in m.groups())
        dados.append([cta, descr, sal_ob, saldo_ob, periodo, saldo_cb])

    cols = ["CTA", "Descripción", "Sal OB", "Saldo OB", "Período", "Saldo CB"]
    df = pd.DataFrame(dados, columns=cols)

    for c in ["Sal OB", "Saldo OB", "Período", "Saldo CB"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


# -----------------------------------------------------------------------------
# PARSER GL0061 — colunas fixas
# -----------------------------------------------------------------------------

def parse_cuenta_gl(texto: str) -> pd.DataFrame:
    """
    Parser para arquivos GL0061 (colunas fixas em linha), com:
    CTA | CC | PROD | CNT | TDW | Fecha | Transacción | Debe | Haber | Saldo Real | Saldo | Texto

    Ajustes:
    - Debe/Haber aceitam número com negativo no final (ex.: "9,200.29-") e preservam o sinal.
    """
    linhas = texto.splitlines()
    dados = []

    cta_header = None
    reg_header = re.compile(r"Nº de cta\.\s+(\d{6})")
    for ln in linhas[:50]:
        m = reg_header.search(ln)
        if m:
            cta_header = m.group(1)
            break
    if not cta_header:
        raise ValueError("CTA não encontrada no cabeçalho do arquivo GL0061.")

    def clean_num(v: str | None) -> float:
        if not v:
            return 0.0
        s = str(v).strip()
        neg = s.endswith("-")
        if neg:
            s = s[:-1]
        s = s.replace(",", "")
        try:
            val = float(s)
        except Exception:
            val = 0.0
        return -val if neg else val

    ignore = re.compile(
        r"Electrolux|Planificación|Moneda|Scala|^-{3,}|^={3,}|"
        r"Saldo Inicial|Saldo final|T O T A L|ACTIVO|Página|Criterios|CUENTAS POR"
    )

    cols = [
        "CTA", "CC", "PROD", "CNT", "TDW",
        "Fecha", "Transacción",
        "Debe", "Haber",
        "Saldo Real", "Saldo",
        "Texto"
    ]

    for ln in linhas:
        if ignore.search(ln):
            continue
        if len(ln.strip()) == 0:
            continue
        if not re.search(r"\d{2}/\d{2}/\d{2}", ln):
            continue

        cc = ln[0:5].strip()
        prod = ln[5:13].strip()
        cnt = ln[13:23].strip()
        tdw = ln[23:31].strip()
        fecha = ln[31:40].strip()
        ntran = ln[40:50].strip()

        nums = re.findall(r"[-\d,]+\.\d{2}-?", ln)
        if len(nums) < 3:
            continue

        debe = clean_num(nums[-3])
        haber = clean_num(nums[-2])
        saldo_impresso = clean_num(nums[-1])

        saldo_real = round(debe - haber, 2)

        texto_pos = ln.rfind(nums[-1])
        texto = ln[texto_pos + len(nums[-1]):].strip() if texto_pos != -1 else ""

        if not cc.isdigit():
            cc = ""

        dados.append([
            cta_header, cc, prod, cnt, tdw,
            fecha, ntran, debe, haber,
            saldo_real, saldo_impresso,
            texto
        ])

    df = pd.DataFrame(dados, columns=cols)

    for date_col in ["Fecha", "Fechado"]:
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True).dt.date

    return df


# -----------------------------------------------------------------------------
# Export XLSX com máscara numérica e data
# -----------------------------------------------------------------------------

def to_xlsx_bytes_format(
    df: pd.DataFrame,
    sheet_name: str,
    numeric_cols: list[str] | None = None,
    date_cols: list[str] | None = None
) -> bytes:
    """
    Exporta para XLSX aplicando:
      - número: #,##0.00
      - data: dd/mm/yyyy
    """
    numeric_cols = numeric_cols or []
    date_cols = date_cols or []

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_to_save = df.copy()
        for dc in date_cols:
            if dc in df_to_save.columns:
                df_to_save[dc] = pd.to_datetime(df_to_save[dc], errors="coerce")

        df_to_save.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        BLUE = "FF0077B6"
        WHITE = "FFFFFFFF"
        fill_blue = PatternFill(fill_type="solid", start_color=BLUE, end_color=BLUE)
        font_white_bold = Font(color=WHITE, bold=True)
        for cell in ws[1]:
            cell.fill = fill_blue
            cell.font = font_white_bold

        for col_name in numeric_cols:
            if col_name not in df_to_save.columns:
                continue
            col_idx = df_to_save.columns.get_loc(col_name) + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.number_format = '#,##0.00'

        for col_name in date_cols:
            if col_name not in df_to_save.columns:
                continue
            col_idx = df_to_save.columns.get_loc(col_name) + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    cell.number_format = 'dd/mm/yyyy'

        for col_idx in range(1, ws.max_column + 1):
            max_len = 10
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col_idx).value
                if v is None:
                    continue
                if isinstance(v, (int, float)):
                    s = f"{v:,.2f}"
                else:
                    s = str(v)
                max_len = max(max_len, len(s))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buffer.seek(0)
    return buffer.getvalue()


# -----------------------------------------------------------------------------
# Página
# -----------------------------------------------------------------------------

def render():
    _ensure_state()

    # --- MIGRAÇÃO: se ainda existir a chave antiga, preserve como "orig" ---
    if "aag_plantilla_df_orig" not in st.session_state and "aag_plantilla_df" in st.session_state:
        try:
            st.session_state["aag_plantilla_df_orig"] = st.session_state["aag_plantilla_df"].copy()
        except Exception:
            st.session_state["aag_plantilla_df_orig"] = st.session_state["aag_plantilla_df"]
        del st.session_state["aag_plantilla_df"]

    st.subheader("Aplicación Archivo Gastos")

    col_b1, col_b2, col_b3, col_b4, col_b5 = st.columns(5)

    with col_b1:
        if st.button("Estado de Cuenta", use_container_width=True):
            _set_mode("estado")
            st.session_state["aag_state"]["last_action"] = "estado"
            st.rerun()
    with col_b2:
        if st.button("Plantilla Gastos", use_container_width=True):
            _set_mode("plantilla")
            st.session_state["aag_state"]["last_action"] = "plantilla"
            st.rerun()
    with col_b3:
        if st.button("Analise", use_container_width=True):
            _set_mode("asientos")
            st.session_state["aag_state"]["last_action"] = "asientos"
            st.rerun()
    with col_b4:
        if st.button("Cuenta", use_container_width=True):
            _set_mode("cuenta")
            st.session_state["aag_state"]["last_action"] = "cuenta"
            st.rerun()
    with col_b5:
        if st.button("Limpieza Plantilla Gastos", use_container_width=True):
            _set_mode("limpieza")  # <<< vira um modo persistente
            st.session_state["aag_state"]["last_action"] = "limpieza_pg"
            st.rerun()

    mode = st.session_state["aag_mode"]
    st.divider()

    # ====== MODO: Limpieza Plantilla Gastos ======
    if mode == "limpieza":
        st.subheader("🧹 Limpieza da Plantilla de Gastos")
        try:
            df_pg_orig = st.session_state.get("aag_plantilla_df_orig", None)
            df_ct = st.session_state.get("aag_cuenta_df", None)

            if df_pg_orig is None or df_pg_orig.empty:
                st.error("Antes de limpar, carregue e execute a **Plantilla de Gastos**.")
            elif df_ct is None or df_ct.empty:
                st.error("Antes de limpar, carregue e processe o **Archivo de Cuenta (GL0061)**.")
            else:
                df_pg_clean, stats = limpiar_plantilla_contra_cuenta(df_pg_orig, df_ct, chave_col="Chave")
                st.session_state["aag_plantilla_df_clean"] = df_pg_clean.copy()
                st.session_state["aag_state"]["last_action"] = "limpieza_pg"

                c1, c2, c3 = st.columns(3)
                with c1:
                    st.metric("Linhas (Original)", f"{stats['rows_original']:,}".replace(",", "."))
                with c2:
                    st.metric("Linhas (Limpo)", f"{stats['rows_clean']:,}".replace(",", "."))
                with c3:
                    st.metric("Removidas", f"{stats['rows_removed']:,}".replace(",", "."))

                if stats["rows_removed"] == 0:
                    st.info("Nenhuma divergência de contagem encontrada. Nada foi removido.")

                # ... depois de montar df_cmp e antes do bloco de métricas existentes:
    
                # =========================
                # Analise (pós-limpeza)
                # =========================
                st.divider()
                st.subheader("🔍 Analise (Estado de Cuenta × Plantilla **limpa**)")

                # Precisamos do Estado e da Plantilla limpa
                df_ec = st.session_state.get("aag_estado_df", None)
                df_pg_clean = st.session_state.get("aag_plantilla_df_clean", None)
                df_pg_orig = st.session_state.get("aag_plantilla_df_orig", None)

                if df_ec is None or df_ec.empty:
                    st.warning("Para rodar a análise, primeiro carregue e execute o **Estado de Cuenta**.")
                elif df_pg_clean is None or df_pg_clean.empty:
                    st.warning("Nenhuma Plantilla limpa encontrada. Execute a **Limpieza Plantilla Gastos**.")
                else:
                    # Tolerância de comparação (igual à do botão Analise)
                    tol = st.number_input("Valor de Tolerância", min_value=0.00, value=0.01, step=0.01, key="limpieza_tol")

                    def _norm_conta(x) -> str:
                        s = re.sub(r"\D", "", str(x))
                        s = s.lstrip("0")
                        return s if s else ""

                    # --- Consolida Estado de Cuenta (CTA x Período) ---
                    try:
                        if "CTA" not in df_ec.columns or "Período" not in df_ec.columns:
                            st.error("Estado de Cuenta não contém as colunas esperadas: 'CTA' e 'Período'.")
                            return  

                        df_ec_proc = df_ec.copy()
                        df_ec_proc["CTA"] = df_ec_proc["CTA"].apply(_norm_conta)
                        df_ec_proc["Período"] = pd.to_numeric(df_ec_proc["Período"], errors="coerce").fillna(0.0)
                        df_ec_proc = df_ec_proc[df_ec_proc["CTA"].astype(str).str.len() > 0]

                        df_ec_agg = (
                            df_ec_proc.groupby("CTA", as_index=False)["Período"]
                            .sum()
                            .rename(columns={"CTA": "Cuenta", "Período": "Saldo_Estado_Cuenta"})
                        )
                    except Exception as e:
                        st.error("Erro ao consolidar Estado de Cuenta (pós-limpeza).")
                        st.exception(e)
                        st.stop()

                    # Helper para localizar colunas (case-insensitive, substrings)
                    def _find_col(df: pd.DataFrame, target: str):
                        for c in df.columns:
                            if str(c).strip().lower() == target:
                                return c
                        cand = [c for c in df.columns if target in str(c).strip().lower()]
                        return cand[0] if cand else None

                    # --- Consolida Plantilla LIMPA (Cuenta x Amount) ---
                    try:
                        cuenta_col_clean = _find_col(df_pg_clean, "cuenta")
                        amount_col_clean = _find_col(df_pg_clean, "amount")

                        if cuenta_col_clean is None or amount_col_clean is None:
                            st.error("Plantilla (limpa) não contém as colunas esperadas: 'Cuenta' e 'Amount'.")
                            st.stop()

                        df_pg_c = df_pg_clean.copy()
                        df_pg_c[amount_col_clean] = pd.to_numeric(df_pg_c[amount_col_clean], errors="coerce").fillna(0.0)
                        df_pg_c["__conta__"] = df_pg_c[cuenta_col_clean].apply(_norm_conta)
                        df_pg_c = df_pg_c[df_pg_c["__conta__"].astype(str).str.len() > 0]

                        df_pg_clean_agg = (
                            df_pg_c.groupby("__conta__", as_index=False)[amount_col_clean]
                            .sum()
                            .rename(columns={"__conta__": "Cuenta", amount_col_clean: "Saldo_Plantilla_Gastos"})
                        )
                    except Exception as e:
                        st.error("Erro ao consolidar Plantilla (limpa).")
                        st.exception(e)
                        st.stop()

                    # --- Consolida Plantilla ORIGINAL (para identificar contas ajustadas) ---
                    df_pg_orig_agg = None
                    contas_ajustadas = pd.DataFrame(columns=["Cuenta", "Ajustada"])
                    try:
                        if df_pg_orig is not None and not df_pg_orig.empty:
                            cuenta_col_orig = _find_col(df_pg_orig, "cuenta")
                            amount_col_orig = _find_col(df_pg_orig, "amount")

                            if cuenta_col_orig and amount_col_orig:
                                df_pg_o = df_pg_orig.copy()
                                df_pg_o[amount_col_orig] = pd.to_numeric(df_pg_o[amount_col_orig], errors="coerce").fillna(0.0)
                                df_pg_o["__conta__"] = df_pg_o[cuenta_col_orig].apply(_norm_conta)
                                df_pg_o = df_pg_o[df_pg_o["__conta__"].astype(str).str.len() > 0]

                                df_pg_orig_agg = (
                                    df_pg_o.groupby("__conta__", as_index=False)[amount_col_orig]
                                    .sum()
                                    .rename(columns={"__conta__": "Cuenta", amount_col_orig: "Saldo_Plantilla_Original"})
                                )

                                # Junta original × limpa para detectar se a conta mudou
                                df_cmp_adj = pd.merge(df_pg_orig_agg, df_pg_clean_agg, on="Cuenta", how="outer")
                                for c in ["Saldo_Plantilla_Original", "Saldo_Plantilla_Gastos"]:
                                    df_cmp_adj[c] = pd.to_numeric(df_cmp_adj[c], errors="coerce").fillna(0.0)

                                # Ajustada = saldo mudou após a limpeza (considera tolerância)
                                contas_ajustadas = df_cmp_adj.assign(
                                    Ajustada=(df_cmp_adj["Saldo_Plantilla_Original"] - df_cmp_adj["Saldo_Plantilla_Gastos"]).abs() > float(tol)
                                )[["Cuenta", "Ajustada"]]
                    except Exception as e:
                        st.error("Erro ao calcular contas ajustadas.")
                        st.exception(e)
                        st.stop()

                    # --- Merge final (Estado × Plantilla LIMPA) e cálculo de diferença ---
                    df_cmp = pd.merge(df_ec_agg, df_pg_clean_agg, on="Cuenta", how="outer")
                    for c in ["Saldo_Estado_Cuenta", "Saldo_Plantilla_Gastos"]:
                        df_cmp[c] = pd.to_numeric(df_cmp[c], errors="coerce").fillna(0.0)

                    df_cmp["Diferença"] = (df_cmp["Saldo_Plantilla_Gastos"] - df_cmp["Saldo_Estado_Cuenta"]).round(2)

                    # Marca divergência pela tolerância
                    df_cmp["_div"] = df_cmp["Diferença"].abs() > float(tol)

                    # Junta a flag de Ajustada (se disponível)
                    if not contas_ajustadas.empty:
                        df_cmp = pd.merge(df_cmp, contas_ajustadas, on="Cuenta", how="left")
                        df_cmp["Ajustada"] = df_cmp["Ajustada"].fillna(False)
                    else:
                        df_cmp["Ajustada"] = False

                    # Ordena por número de conta
                    df_cmp["_cuenta_num"] = pd.to_numeric(df_cmp["Cuenta"], errors="coerce")
                    df_cmp = df_cmp.sort_values(by="_cuenta_num", ascending=True).drop(columns=["_cuenta_num"]).reset_index(drop=True)

                    # ---- Filtros ----
                    col_f1, col_f2 = st.columns(2)
                    with col_f1:
                        only_div = st.checkbox("Mostrar apenas contas com divergência", value=True, key="limpieza_only_div")
                    with col_f2:
                        only_adj = st.checkbox("Mostrar apenas contas Ajustadas", value=False, key="limpieza_only_adj")

                    df_show = df_cmp.copy()
                    if only_div:
                        df_show = df_show[df_show["_div"]]
                    if only_adj:
                        df_show = df_show[df_show["Ajustada"]]

                    # Colunas finais para exibição
                    df_show = df_show[["Cuenta", "Saldo_Estado_Cuenta", "Saldo_Plantilla_Gastos", "Diferença", "Ajustada"]]

                    # Exibição
                    st.dataframe(
                        df_show,
                        use_container_width=True, height=520,
                        column_config={
                            "Saldo_Estado_Cuenta": st.column_config.NumberColumn(format="%.2f"),
                            "Saldo_Plantilla_Gastos": st.column_config.NumberColumn(format="%.2f"),
                            "Diferença": st.column_config.NumberColumn(format="%.2f"),
                            "Ajustada": st.column_config.CheckboxColumn(),
                        },
                    )

                    # Downloads
                    col_d1, col_d2 = st.columns(2)
                    with col_d1:
                        st.download_button(
                            label="Baixar CSV (Analise pós-limpeza)",
                            data=df_show.to_csv(index=False).encode("utf-8"),
                            file_name="analise_pos_limpieza.csv",
                            mime="text/csv",
                            use_container_width=True,
                        )
                    with col_d2:
                        xlsx_bytes_cmp = to_xlsx_bytes_format(
                            df_show,
                            sheet_name="AnalisePosLimpieza",
                            numeric_cols=["Saldo_Estado_Cuenta", "Saldo_Plantilla_Gastos", "Diferença"],
                            date_cols=[],
                        )
                        st.download_button(
                            label="Baixar XLSX (Analise pós-limpeza)",
                            data=xlsx_bytes_cmp,
                            file_name="analise_pos_limpieza.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

                # Prévia da Plantilla de Gastos (após limpeza) — SOMENTE nesta aba
                df_pg_prev = st.session_state["aag_plantilla_df_clean"]

                amount_col_view = None
                for c in df_pg_prev.columns:
                    if str(c).strip().lower() == "amount":
                        amount_col_view = c
                        break
                if amount_col_view is None:
                    candidates = [c for c in df_pg_prev.columns if "amount" in str(c).strip().lower()]
                    if candidates:
                        amount_col_view = candidates[0]

                known_date_keys = {"transactiondate", "duedate", "due_date", "invoicedate", "invoice_date"}
                found_date_cols_view = [c for c in df_pg_prev.columns if c.lower().replace(" ", "_") in known_date_keys]

                col_cfg = {}
                if amount_col_view:
                    col_cfg[str(amount_col_view)] = st.column_config.NumberColumn(format="%.2f")
                for dc in found_date_cols_view:
                    if pd.api.types.is_datetime64_any_dtype(df_pg_prev[dc]):
                        col_cfg[dc] = st.column_config.DateColumn(format="DD/MM/YYYY")
                    else:
                        col_cfg[dc] = st.column_config.TextColumn()

                # st.dataframe(df_pg_prev, use_container_width=True, height=520, column_config=col_cfg)

                col_csv, col_xlsx = st.columns(2)
                with col_csv:
                    st.download_button(
                        "Baixar CSV (Plantilla Limpia)",
                        df_pg_prev.to_csv(index=False).encode("utf-8"),
                        "plantilla_gastos_limpia.csv",
                        "text/csv",
                        use_container_width=True
                    )

        except Exception as e:
            st.error("Erro durante a limpeza da Plantilla de Gastos.")
            st.exception(e)

        return  # encerra o render no modo 'limpieza'

    # -------------------------------------------------------------------------
    # Modo: Estado de Cuenta (.txt)
    # -------------------------------------------------------------------------
    if mode == "estado":
        upl_key_estado = st.session_state["aag_state"].setdefault("uploader_key_estado", "aag_estado_upl_1")

        st.caption("Carregue o arquivo **.txt** de *Listado de Saldos* para visualização e export.")
        uploaded = st.file_uploader(
            "Selecionar arquivo (.txt)",
            type=["txt"],
            accept_multiple_files=False,
            key=upl_key_estado,
            help="Ex.: relatório 'Listado de Saldos' exportado do sistema.",
        )

        col_run, col_clear = st.columns([2, 1])
        with col_run:
            run_clicked = st.button("▶️ Executar", type="primary", use_container_width=True, disabled=(uploaded is None))
        with col_clear:
            clear_clicked = st.button("Limpar", use_container_width=True)

        if clear_clicked:
            st.session_state["aag_state"]["uploader_key_estado"] = upl_key_estado + "_x"
            if "aag_estado_df" in st.session_state:
                del st.session_state["aag_estado_df"]
            st.rerun()

        if run_clicked and uploaded is not None:
            pbar = st.progress(0, text="Lendo arquivo .txt...")
            try:
                raw_bytes = uploaded.getvalue()
                try:
                    text = raw_bytes.decode("utf-8")
                except UnicodeDecodeError:
                    text = raw_bytes.decode("latin-1")

                pbar.progress(35, text="Convertendo para DataFrame...")
                df_base = parse_estado_cuenta_txt(text)

                if df_base is None or df_base.empty:
                    st.warning("Nenhuma linha válida encontrada no arquivo.")
                    pbar.progress(0, text="Aguardando...")
                    return

                st.session_state["aag_estado_df"] = df_base.copy()

                pbar.progress(70, text="Preparando visualização...")
                st.success("Arquivo processado com sucesso.")
                pbar.progress(100, text="Concluído.")
            except Exception as e:
                st.error("Erro ao processar o arquivo .txt.")
                st.exception(e)

        if "aag_estado_df" in st.session_state and isinstance(st.session_state["aag_estado_df"], pd.DataFrame):
            df_base = st.session_state["aag_estado_df"]
        
            # --- KPIs solicitados ---
            # Contagem de contas (CTA distintas)
            contas_distintas = df_base["CTA"].astype(str).str.strip().replace({"": np.nan}).dropna().nunique() if "CTA" in df_base.columns else 0
        
            # Soma da coluna Período (sem incluir a linha TOTAL)
            soma_periodo = 0.0
            if "Período" in df_base.columns:
                soma_periodo = pd.to_numeric(df_base["Período"], errors="coerce").fillna(0.0).sum()
        
            # Exibição dos KPIs
            kpi1, kpi2 = st.columns(2)
            with kpi1:
                st.metric("Contagem de Contas no arquivo", f"{contas_distintas:,}".replace(",", "."))
            with kpi2:
                st.metric(
                    "Soma Coluna Período",
                    f"{soma_periodo:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                )
        
            # --- Tabela com linha TOTAL (como já existia) ---
            df = df_base.copy()
            numeric_cols = ["Sal OB", "Saldo OB", "Período", "Saldo CB"]
            for c in numeric_cols:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors="coerce")
        
            totals = {c: float(np.nansum(df[c].values)) for c in numeric_cols if c in df.columns}
        
            # Adiciona linha TOTAL ao final (sem afetar os KPIs acima)
            total_row = {col: "" for col in df.columns}
            if "Descripción" in df.columns:
                total_row["Descripción"] = "TOTAL"
            for c in numeric_cols:
                if c in totals:
                    total_row[c] = totals[c]
            df = pd.concat([df, pd.DataFrame([total_row], columns=df.columns)], ignore_index=True)
        
            st.dataframe(
                df,
                use_container_width=True,
                height=550,
                column_config={c: st.column_config.NumberColumn(format="%.2f") for c in numeric_cols if c in df.columns},
            )
        
            col_csv, col_xlsx = st.columns(2)
            with col_csv:
                st.download_button(
                    label="Baixar CSV (Estado de Cuenta)",
                    data=df.to_csv(index=False).encode("utf-8"),
                    file_name="estado_de_cuenta.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with col_xlsx:
                xlsx_bytes = to_xlsx_bytes_format(
                    df, sheet_name="EstadoCuenta", numeric_cols=[c for c in numeric_cols if c in df.columns], date_cols=[]
                )
                st.download_button(
                    label="Baixar XLSX (Estado de Cuenta)",
                    data=xlsx_bytes,
                    file_name="estado_de_cuenta.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    # -------------------------------------------------------------------------
    # Modo: Plantilla de Gastos (.xlsx/.xls)
    # -------------------------------------------------------------------------
    elif mode == "plantilla":
        upl_key_pg = st.session_state["aag_state"].setdefault("uploader_key_pg", "aag_pg_upl_1")

        st.caption("Carregue o arquivo **Excel** da *Plantilla de Gastos* (primeira aba será lida).")
        uploaded_xl = st.file_uploader(
            "Selecionar arquivo (.xlsx ou .xls)",
            type=["xlsx", "xls"],
            accept_multiple_files=False,
            key=upl_key_pg,
            help="As colunas de data serão convertidas para dd/mm/aaaa e Amount formatado como número.",
        )

        col_run, col_clear = st.columns([2, 1])
        with col_run:
            run_clicked = st.button("▶️ Executar", type="primary", use_container_width=True, disabled=(uploaded_xl is None))
        with col_clear:
            clear_clicked = st.button("Limpar", use_container_width=True)

        if clear_clicked:
            st.session_state["aag_state"]["uploader_key_pg"] = upl_key_pg + "_x"
            for k in ("aag_plantilla_df_orig", "aag_plantilla_df_clean"):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()

        if run_clicked and uploaded_xl is not None:
            pbar = st.progress(0, text="Lendo arquivo Excel...")
            try:
                name = getattr(uploaded_xl, "name", "").lower()
                engine = "openpyxl" if name.endswith(".xlsx") else "xlrd"

                # Lê como texto para preservar zeros à esquerda (Amount e datas tratadas depois)
                df_pg = pd.read_excel(uploaded_xl, sheet_name=0, engine=engine, dtype=str)

                # --- Sanitização: remover quebras de linha de todas as colunas texto ---
                for col in df_pg.columns:
                    # Somente colunas de texto; Amount será convertido depois
                    if df_pg[col].dtype == object:
                        df_pg[col] = _remove_newlines(df_pg[col])
                
                # Se quiser garantir especificamente para a coluna PPQ:
                # (caso o nome varie, ajuste conforme seu arquivo)
                possible_ppq_cols = [c for c in df_pg.columns if str(c).strip().lower() in {"ppq"}]
                for c in possible_ppq_cols:
                    df_pg[c] = _remove_newlines(df_pg[c])
                    
                # Detecta coluna Amount (case-insensitive)
                amount_col = None
                for c in df_pg.columns:
                    if str(c).strip().lower() == "amount":
                        amount_col = c
                        break
                if amount_col is None:
                    candidates = [c for c in df_pg.columns if "amount" in str(c).strip().lower()]
                    if candidates:
                        amount_col = candidates[0]
                if amount_col is None:
                    st.error("Coluna 'Amount' não encontrada no arquivo.")
                    return

                # Amount numérico
                df_pg[amount_col] = pd.to_numeric(df_pg[amount_col], errors="coerce")

                # --- Normalização de datas APENAS na Plantilla (forçar dd/mm/yyyy na Chave) ---
                def norm(s: str) -> str:
                    return re.sub(r"[^a-z0-9]", "", str(s).strip().lower())

                # Localiza colunas alvo
                date_targets = {"transactiondate": None, "duedate": None, "invoicedate": None}
                for c in df_pg.columns:
                    nc = norm(c)
                    if nc == "transactiondate":
                        date_targets["transactiondate"] = c
                    elif nc in ("duedate", "due_date"):
                        date_targets["duedate"] = c
                    elif nc in ("invoicedate", "invoice_date"):
                        date_targets["invoicedate"] = c

                found_date_cols = []
                for key, col in date_targets.items():
                    if col and col in df_pg.columns:
                        # Converte qualquer mistura (serial Excel, '2026-01-02 00:00:00', etc.) -> datetime64[ns]
                        df_pg[col] = _to_datetime_from_mixed_excel_and_strings(df_pg[col])
                        found_date_cols.append(col)

                # Helper para achar colunas por variações de nome
                def _find_col_ci(df: pd.DataFrame, targets: list[str]):
                    cols_map = {re.sub(r"[^a-z0-9]", "", str(c).lower()): c for c in df.columns}
                    for t in targets:
                        key = re.sub(r"[^a-z0-9]", "", t.lower())
                        if key in cols_map:
                            return cols_map[key]
                    return None

                cuenta_col = _find_col_ci(df_pg, ["Cuenta"])
                tno_col = _find_col_ci(df_pg, ["TransactionNo", "Transaction No", "TransNo", "Transaction_Number"])
                amount_col_ci = amount_col

                # Formata data da CHAVE explicitamente em dd/mm/yyyy
                tdate_col = date_targets.get("transactiondate")
                tdate_str = _fmt_date_series_ddmmyyyy(df_pg[tdate_col]) if tdate_col else pd.Series([""] * len(df_pg))

                tno_str = df_pg[tno_col].apply(_fmt_transno_keep_zeros) if tno_col else ""
                cuenta_str = df_pg[cuenta_col].apply(_str_or_empty) if cuenta_col else ""
                amount_str = df_pg[amount_col_ci].apply(_fmt_num_2dec_point) if amount_col_ci else ""

                def _ensure_series(x, n):
                    return x if isinstance(x, pd.Series) else pd.Series([""] * n)

                df_pg["Chave"] = (
                    _ensure_series(cuenta_str, len(df_pg)) + "|" +
                    _ensure_series(tdate_str, len(df_pg)) + "|" +
                    _ensure_series(tno_str, len(df_pg)) + "|" +
                    _ensure_series(amount_str, len(df_pg))
                )

                st.session_state["aag_plantilla_df_orig"] = df_pg.copy()

                pbar.progress(70, text="Preparando visualização...")
                st.success("Arquivo carregado com sucesso.")
                pbar.progress(100, text="Concluído.")
            except Exception as e:
                st.error("Erro ao processar o arquivo Excel.")
                st.exception(e)

        # Exibição e downloads
        if "aag_plantilla_df_orig" in st.session_state and isinstance(st.session_state["aag_plantilla_df_orig"], pd.DataFrame):
            df_pg = st.session_state["aag_plantilla_df_orig"]

            amount_col_view = None
            for c in df_pg.columns:
                if str(c).strip().lower() == "amount":
                    amount_col_view = c
                    break
            if amount_col_view is None:
                candidates = [c for c in df_pg.columns if "amount" in str(c).strip().lower()]
                if candidates:
                    amount_col_view = candidates[0]

            known_date_keys = {"transactiondate", "duedate", "due_date", "invoicedate", "invoice_date"}
            found_date_cols_view = [c for c in df_pg.columns if c.lower().replace(" ", "_") in known_date_keys]

            col_cfg = {}
            if amount_col_view:
                col_cfg[str(amount_col_view)] = st.column_config.NumberColumn(format="%.2f")
            for dc in found_date_cols_view:
                if pd.api.types.is_datetime64_any_dtype(df_pg[dc]):
                    col_cfg[dc] = st.column_config.DateColumn(format="DD/MM/YYYY")
                else:
                    col_cfg[dc] = st.column_config.TextColumn()

            # st.dataframe(df_pg, use_container_width=True, height=550, column_config=col_cfg)

            col_csv, col_xlsx = st.columns(2)
            with col_csv:
                st.download_button(
                    label="Baixar CSV (Plantilla Gastos)",
                    data=df_pg.to_csv(index=False).encode("utf-8"),
                    file_name="plantilla_gastos.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

            if "aag_plantilla_df_clean" in st.session_state:
                st.caption("Uma versão **limpa** foi gerada. Para baixá-la, use o botão **Limpieza Plantilla Gastos**.")

    # -------------------------------------------------------------------------
    # Modo: Analise — compara Estado de Cuenta (CTA/Período) x Plantilla (Cuenta/Amount)
    # -------------------------------------------------------------------------
    elif mode == "asientos":
        st.subheader("🔍 Analise: Estado de Cuenta x Plantilla de Gastos")

        df_ec = st.session_state.get("aag_estado_df", None)
        df_pg = st.session_state.get("aag_plantilla_df_orig", None)

        missing = []
        if df_ec is None or df_ec.empty:
            missing.append("Estado de Cuenta (.txt)")
        if df_pg is None or df_pg.empty:
            missing.append("Plantilla de Gastos (.xlsx/.xls)")

        if missing:
            st.warning(
                "Para executar a análise, primeiro carregue: " + ", ".join(missing) +
                ". Use as abas **Estado de Cuenta** e **Plantilla Gastos**."
            )
            return

        tol = st.number_input("Valor de Tolerância", min_value=0.00, value=0.01, step=0.01)

        def _norm_conta(x) -> str:
            s = re.sub(r"\D", "", str(x))
            s = s.lstrip("0")
            return s if s else ""

        pbar = st.progress(0, text="Consolidando Estado de Cuenta...")

        try:
            if "CTA" not in df_ec.columns or "Período" not in df_ec.columns:
                st.error("Estado de Cuenta não contém as colunas esperadas: 'CTA' e 'Período'.")
                return

            df_ec_proc = df_ec.copy()
            df_ec_proc["CTA"] = df_ec_proc["CTA"].apply(_norm_conta)
            df_ec_proc["Período"] = pd.to_numeric(df_ec_proc["Período"], errors="coerce").fillna(0.0)
            df_ec_proc = df_ec_proc[df_ec_proc["CTA"].astype(str).str.len() > 0]

            df_ec_agg = (
                df_ec_proc.groupby("CTA", as_index=False)["Período"]
                .sum()
                .rename(columns={"CTA": "Cuenta", "Período": "Saldo_Estado_Cuenta"})
            )

            pbar.progress(40, text="Consolidando Plantilla de Gastos...")
        except Exception as e:
            st.error("Erro ao consolidar Estado de Cuenta.")
            st.exception(e)
            return

        try:
            def _find_col(df, target):
                for c in df.columns:
                    if str(c).strip().lower() == target:
                        return c
                cand = [c for c in df.columns if target in str(c).strip().lower()]
                return cand[0] if cand else None

            cuenta_col = _find_col(df_pg, "cuenta")
            amount_col = _find_col(df_pg, "amount")

            if cuenta_col is None or amount_col is None:
                st.error("Plantilla não contém as colunas esperadas: 'Cuenta' e 'Amount'.")
                return

            df_pg_proc = df_pg.copy()
            df_pg_proc[amount_col] = pd.to_numeric(df_pg_proc[amount_col], errors="coerce").fillna(0.0)
            df_pg_proc["__conta__"] = df_pg_proc[cuenta_col].apply(_norm_conta)
            df_pg_proc = df_pg_proc[df_pg_proc["__conta__"].astype(str).str.len() > 0]

            df_pg_agg = (
                df_pg_proc.groupby("__conta__", as_index=False)[amount_col]
                .sum()
                .rename(columns={"__conta__": "Cuenta", amount_col: "Saldo_Plantilla_Gastos"})
            )

            pbar.progress(70, text="Comparando saldos...")
        except Exception as e:
            st.error("Erro ao consolidar Plantilla de Gastos.")
            st.exception(e)
            return

        try:
            df_cmp = pd.merge(df_ec_agg, df_pg_agg, on="Cuenta", how="outer")
            for c in ["Saldo_Estado_Cuenta", "Saldo_Plantilla_Gastos"]:
                df_cmp[c] = pd.to_numeric(df_cmp[c], errors="coerce").fillna(0.0)

            df_cmp["Diferença"] = (df_cmp["Saldo_Plantilla_Gastos"] - df_cmp["Saldo_Estado_Cuenta"]).round(2)

            df_cmp["_div"] = df_cmp["Diferença"].abs() > float(tol)

            df_cmp["_cuenta_num"] = pd.to_numeric(df_cmp["Cuenta"], errors="coerce")
            df_cmp = df_cmp.sort_values(by="_cuenta_num", ascending=True).drop(columns=["_cuenta_num"]).reset_index(drop=True)

            pbar.progress(90, text="Preparando visualização...")

            # ... depois de montar df_cmp e antes do bloco de métricas existentes:
            
            # Contagem de contas com divergência (únicas)
            div_count = int(df_cmp.loc[df_cmp["_div"], "Cuenta"].nunique())
            
            # Você já tem:
            # c1, c2, c3 = st.columns(3)
            # Vamos aumentar para 4 colunas e incluir a diverência:
            
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.metric("Contas (Estado)", f"{df_ec_agg['Cuenta'].nunique():,}".replace(",", "."))
            with c2:
                st.metric(
                    "Soma Estado de Cuentas",
                    f"{df_ec_agg['Saldo_Estado_Cuenta'].sum():,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                )
            with c3:
                st.metric(
                    "Soma Plantilla de Gastos",
                    f"{df_pg_agg['Saldo_Plantilla_Gastos'].sum():,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                )
            with c4:
                # Título em espanhol, seguindo o pedido
                st.metric("Cuentas con divergencia", f"{div_count:,}".replace(",", "."))

            only_div = st.checkbox("Mostrar apenas contas com divergência", value=True)
            df_show = df_cmp[df_cmp["_div"]] if only_div else df_cmp

            df_show = df_show[["Cuenta", "Saldo_Estado_Cuenta", "Saldo_Plantilla_Gastos", "Diferença"]]

            st.dataframe(
                df_show,
                use_container_width=True, height=520,
                column_config={
                    "Saldo_Estado_Cuenta": st.column_config.NumberColumn(format="%.2f"),
                    "Saldo_Plantilla_Gastos": st.column_config.NumberColumn(format="%.2f"),
                    "Diferença": st.column_config.NumberColumn(format="%.2f"),
                },
            )

            pbar.progress(95, text="Gerando arquivos para download...")

            col_d1, col_d2 = st.columns(2)
            with col_d1:
                st.download_button(
                    label="Baixar CSV (Analise)",
                    data=df_show.to_csv(index=False).encode("utf-8"),
                    file_name="analise_contas.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with col_d2:
                xlsx_bytes = to_xlsx_bytes_format(
                    df_show,
                    sheet_name="Analise",
                    numeric_cols=["Saldo_Estado_Cuenta", "Saldo_Plantilla_Gastos", "Diferença"],
                    date_cols=[],
                )
                st.download_button(
                    label="Baixar XLSX (Analise)",
                    data=xlsx_bytes,
                    file_name="analise_contas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            pbar.progress(100, text="Concluído.")

        except Exception as e:
            st.error("Erro durante a comparação.")
            st.exception(e)

    # -------------------------------------------------------------------------
    # Modo: Cuenta (GL0061)
    # -------------------------------------------------------------------------
    elif mode == "cuenta":
        st.subheader("📘 Importar Archivo de Cuenta (GL0061)")

        upl_key = st.session_state["aag_state"].setdefault("uploader_key_cuenta", "aag_cuenta_upl_1")
        uploaded = st.file_uploader("Selecionar arquivo GL0061 (.txt)", type=["txt"], key=upl_key,accept_multiple_files=True)

        col_r, col_c = st.columns([2, 1])
        with col_r:
            run_clicked = st.button("▶️ Processar Cuenta", type="primary", use_container_width=True, disabled=(uploaded is None))
        with col_c:
            clear_clicked = st.button("Limpar", use_container_width=True)

        if clear_clicked:
            st.session_state["aag_state"]["uploader_key_cuenta"] = upl_key + "_x"
            if "aag_cuenta_df" in st.session_state:
                del st.session_state["aag_cuenta_df"]
            st.rerun()

        if run_clicked and uploaded:
            dfs = []
            for up in uploaded:
                try:
                    raw = up.getvalue()
                    try:
                        text = raw.decode("utf-8")
                    except Exception:
                        text = raw.decode("latin-1")
        
                    # Parse de cada arquivo
                    try:
                        df_part = parse_cuenta_gl(text)
                    except Exception as e:
                        st.error(f"Erro ao interpretar o arquivo GL0061: {getattr(up, 'name', '(sem nome)')}")
                        st.exception(e)
                        continue
        
                    if df_part is None or df_part.empty:
                        st.error(f"Nenhuma linha reconhecida no arquivo GL0061: {getattr(up, 'name', '(sem nome)')}")
                        continue
        
                    # Montagem da Chave no MESMO formato que você já usa (com '|')
                    cta_str   = df_part["CTA"].apply(_str_or_empty) if "CTA" in df_part.columns else pd.Series([""] * len(df_part))
                    fecha_str = df_part["Fecha"].apply(_fmt_date_ddmmyyyy) if "Fecha" in df_part.columns else pd.Series([""] * len(df_part))
                    tran_str  = df_part["Transacción"].apply(_fmt_transno_keep_zeros) if "Transacción" in df_part.columns else pd.Series([""] * len(df_part))
                    sreal_str = df_part["Saldo Real"].apply(_fmt_num_2dec_point) if "Saldo Real" in df_part.columns else pd.Series([""] * len(df_part))
        
                    df_part["Chave"] = cta_str + "|" + fecha_str + "|" + tran_str + "|" + sreal_str
        
                    # (Opcional) coluna para rastrear arquivo de origem
                    df_part["Arquivo"] = getattr(up, "name", "")
        
                    dfs.append(df_part)
        
                except Exception as e:
                    st.error(f"Falha ao processar o arquivo: {getattr(up, 'name', '(sem nome)')}")
                    st.exception(e)
        
            if not dfs:
                st.error("Nenhum arquivo GL0061 válido foi processado.")
                return
        
            # Concatena tudo
            df_all = pd.concat(dfs, ignore_index=True)
        
            # (Opcional) se quiser remover duplicatas pelo identificador composto:
            # df_all = df_all.drop_duplicates(subset=["Chave"])
        
            st.session_state["aag_cuenta_df"] = df_all.copy()
            
        if "aag_cuenta_df" in st.session_state and isinstance(st.session_state["aag_cuenta_df"], pd.DataFrame):
            df = st.session_state["aag_cuenta_df"]

            date_cols = [c for c in ["Fecha", "Fechado"] if c in df.columns]
            col_cfg = {
                "Debe": st.column_config.NumberColumn(format="%.2f"),
                "Haber": st.column_config.NumberColumn(format="%.2f"),
                "Saldo Real": st.column_config.NumberColumn(format="%.2f"),
                "Saldo": st.column_config.NumberColumn(format="%.2f"),
            }
            for dc in date_cols:
                col_cfg[dc] = st.column_config.DateColumn(format="DD/MM/YYYY")

            st.dataframe(df, use_container_width=True, height=600, column_config=col_cfg)

            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    "Baixar CSV",
                    df.to_csv(index=False).encode("utf-8"),
                    "cuenta.csv",
                    "text/csv",
                    use_container_width=True
                )
            with col2:
                xlsx_bytes = to_xlsx_bytes_format(
                    df, "Cuenta",
                    numeric_cols=["Debe", "Haber", "Saldo Real", "Saldo"],
                    date_cols=date_cols
                )
                st.download_button(
                    "Baixar XLSX",
                    xlsx_bytes,
                    "cuenta.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    else:
        st.info("Selecione um modo acima para continuar.")
