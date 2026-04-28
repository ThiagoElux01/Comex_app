# ui/pages/process_pdfs.py

import streamlit as st
from io import BytesIO
import pandas as pd

from services import pdf_service
from services.tasa_service import atualizar_dataframe_tasa
from ui.pages import downloads_page

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font


# =========================================================
# Imports protegidos — Adicionales
# =========================================================

ADICIONALES_AVAILABLE = True
ADICIONALES_ERR = None

try:
    from services.adicionales_service import process_adicionales_streamlit
except Exception as e:
    ADICIONALES_AVAILABLE = False
    ADICIONALES_ERR = e


# =========================================================
# Imports protegidos — Percepciones
# =========================================================

PERC_AVAILABLE = True
PERC_ERR = None

try:
    from services.percepcion_service import process_percepcion_streamlit
except Exception as e:
    PERC_AVAILABLE = False
    PERC_ERR = e


# =========================================================
# Imports protegidos — DUAS
# =========================================================

DUAS_AVAILABLE = True
DUAS_ERR = None

try:
    from services.duas_service import process_duas_streamlit
except Exception as e:
    DUAS_AVAILABLE = False
    DUAS_ERR = e


# =========================================================
# Imports protegidos — Externos
# =========================================================

EXTERNOS_AVAILABLE = True
EXTERNOS_ERR = None

try:
    from services.externos_service import process_externos_streamlit
except Exception as e:
    EXTERNOS_AVAILABLE = False
    EXTERNOS_ERR = e


# =========================================================
# Configuração de layout XLSX
# =========================================================

USE_PRN_WIDTHS = True

PRN_WIDTHS_1 = [
    10, 25, 6, 6, 6, 16, 16, 2, 5, 16, 3, 2,
    30, 6, 3, 3, 8, 3, 6, 4, 16, 16, 3, 6
]

PRN_WIDTHS_2 = [
    6, 3, 3, 8, 3, 16, 16, 2, 30, 6, 15, 20, 5
]


def set_fixed_widths(ws, widths, start_col: int = 1):
    for i, w in enumerate(widths, start=start_col):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = float(w) + 0.71


def autofit_worksheet(ws, min_width=8, max_width=60):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(
            max(max_len * 1.2, min_width),
            max_width
        )


def header_paint(ws):
    blue = PatternFill("solid", fgColor="0077B6")
    white_bold = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = blue
        cell.font = white_bold


def to_xlsx_bytes(df: pd.DataFrame, sheet_name="Data") -> bytes:
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        if USE_PRN_WIDTHS:
            if df.shape[1] == 24:
                set_fixed_widths(ws, PRN_WIDTHS_1)
            elif df.shape[1] == 13:
                set_fixed_widths(ws, PRN_WIDTHS_2)
            else:
                autofit_worksheet(ws)
        else:
            autofit_worksheet(ws)

        header_paint(ws)

    buffer.seek(0)
    return buffer.getvalue()


# =========================================================
# Estado
# =========================================================

def _ensure_state():
    if "acao_selecionada" not in st.session_state:
        st.session_state.acao_selecionada = None

    if "uploader_key" not in st.session_state:
        st.session_state.uploader_key = "uploader_none"

    if "tasa_df" not in st.session_state:
        st.session_state.tasa_df = None


def _select_action(action_key: str):
    st.session_state.acao_selecionada = action_key
    st.session_state.uploader_key = f"uploader_{action_key}"


# =========================================================
# Página principal
# =========================================================

def render():
    _ensure_state()

    st.subheader("Aplicación Comex")

    # ---------- Diagnóstico de módulos ----------
    if not ADICIONALES_AVAILABLE:
        st.warning("Módulo **Gastos Adicionales** não pôde ser carregado.")
        with st.expander("Detalhes técnicos (Adicionales)"):
            st.exception(ADICIONALES_ERR)

    if not PERC_AVAILABLE:
        st.warning("Módulo **Percepciones** não pôde ser carregado.")
        with st.expander("Detalhes técnicos (Percepciones)"):
            st.exception(PERC_ERR)

    if not DUAS_AVAILABLE:
        st.warning("Módulo **DUAS** não pôde ser carregado.")
        with st.expander("Detalhes técnicos (DUAS)"):
            st.exception(DUAS_ERR)

    if not EXTERNOS_AVAILABLE:
        st.warning("Módulo **Externos** não pôde ser carregado.")
        with st.expander("Detalhes técnicos (Externos)"):
            st.exception(EXTERNOS_ERR)

    st.divider()

    # ---------- Seleção ----------
    cols = st.columns(4)

    with cols[0]:
        if st.button("Externos"):
            _select_action("externos")

    with cols[1]:
        if st.button("Gastos Adicionales"):
            _select_action("gastos")

    with cols[2]:
        if st.button("DUAS"):
            _select_action("duas")

    with cols[3]:
        if st.button("Percepciones"):
            _select_action("percepciones")

    # ---------- Conteúdo ----------
    action = st.session_state.acao_selecionada

    if not action:
        st.info("Selecione uma operação acima para começar.")
        return

    # ---------- Upload ----------
    uploaded_files = st.file_uploader(
        "Selecione os PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        key=st.session_state.uploader_key,
    )

    # ---------- Tasa ----------
    col1, col2 = st.columns([1, 3])
    with col1:
        if st.button("Atualizar Tasa SUNAT"):
            st.session_state.tasa_df = atualizar_dataframe_tasa()

    with col2:
        if st.session_state.tasa_df is not None:
            st.success("Tasa carregada com sucesso.")

    tasa_df = st.session_state.tasa_df

    # ---------- Processamento ----------
    if uploaded_files:
        with st.spinner("Processando PDFs..."):
            if action == "externos" and EXTERNOS_AVAILABLE:
                df = process_externos_streamlit(uploaded_files, cambio_df=tasa_df)

            elif action == "gastos" and ADICIONALES_AVAILABLE:
                df = process_adicionales_streamlit(uploaded_files, cambio_df=tasa_df)

            elif action == "duas" and DUAS_AVAILABLE:
                df = process_duas_streamlit(uploaded_files, cambio_df=tasa_df)

            elif action == "percepciones" and PERC_AVAILABLE:
                df = process_percepcion_streamlit(uploaded_files)

            else:
                df = None

        if df is not None and not df.empty:
            st.success("Processamento concluído.")
            st.dataframe(df, use_container_width=True)

            xlsx_bytes = to_xlsx_bytes(df)
            st.download_button(
                "📥 Baixar Excel",
                xlsx_bytes,
                file_name=f"{action}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.warning("Nenhum dado foi gerado.")
